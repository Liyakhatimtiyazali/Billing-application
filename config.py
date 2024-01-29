from openpyxl import *
import xlrd
xlrd.xlsx.ensure_elementtree_imported(False, None)
xlrd.xlsx.Element_has_iter = True

def doConnectionwrite():
    try:
        wb=load_workbook('C:\\Users\\lenovo legion 5\\Desktop\\final project\\database\\product.xlsx')
        
        print("File Created******")
        #step2 activate the server
        sheet=wb.active
        return True,sheet,wb
    except:
        print("not connected")
        return False,"null","null"

def doConnectionread():
    try:
        wb=xlrd.open_workbook('C:\\Users\\lenovo legion 5\\Desktop\\final project\\database\\product.xlsx')
        #step2 - activate the database
        sheet=wb.sheet_by_index(0)
       
        return True,sheet
    except:
        print("not connected")
        return False,"null"
    
def docustomerwrite():
    try:
        wb=load_workbook('C:\\Users\\lenovo legion 5\\Desktop\\final project\\database\\paylater.xlsx')
        
        print("File Created******")
        #step2 activate the server
        sheet=wb.active
        return True,sheet,wb
    except:
        print("not connected")
        return False,"null","null"
    
def docustomeread():
    try:
        wb=xlrd.open_workbook('C:\\Users\\lenovo legion 5\\Desktop\\final project\\database\\paylater.xlsx')
        #step2 - activate the database
        sheet=wb.sheet_by_index(0)
       
        return True,sheet
    except:
        print("not connected")
        return False,"null"
    
def dopaymentwrite():
    try:
        wb=load_workbook('C:\\Users\\lenovo legion 5\\Desktop\\final project\\database\\payment.xlsx')
        
        print("File Created******")
        #step2 activate the server
        sheet=wb.active
        return True,sheet,wb
    except:
        print("not connected")
        return False,"null","null"
