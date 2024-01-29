from flask import Flask,request
from openpyxl import *
from connection import config
from service import service,constant
import random


app=Flask(__name__)

cobj= constant.constant()
cart={}
pcart=[]
spcart=[]
mrpcart=[]
qcart=[]
customerbill={}

@app.route('/searchbyname/<sname>')
def searchbyname(sname):
    #sname="nnn"
    status,sheet=config.doConnectionread()
    if(status):
        
        row=sheet.nrows
        print(row)
        #step 4 -retrieve the product details
        productfound=0
        for i in range(row):
            pname=sheet.cell_value(i,0)
            #print(dbuserid)
            if(sname==pname):
                productfound=1
                pname=sheet.cell_value(i,0)
                pid=sheet.cell_value(i,1)
                mrp=sheet.cell_value(i,2)
                sp=sheet.cell_value(i,3)
                quantity=sheet.cell_value(i,4)
                
                return [pname,pid,mrp,sp,quantity]
        if(productfound==0):
            return cobj.impname
    else:
        return cobj.cp
    
@app.route('/searchbyid/<sid>')
def searchbyid(sid):
    #sid="500"
    #step1 -open the database 
    status,sheet=config.doConnectionread()
    if(status):
        #step3 -find out total no of data present in server
        row=sheet.nrows
        print(row)
        #step 4 -retrieve the product details
        productfound=0
        for i in range(row):
            pid=sheet.cell_value(i,1)
            print(pid)
            #print(dbuserid)
            if(sid==pid):
                productfound=1
                pname=sheet.cell_value(i,0)
                pid=sheet.cell_value(i,1)
                mrp=sheet.cell_value(i,2)
                sp=sheet.cell_value(i,3)
                quantity=sheet.cell_value(i,4)
                return [pname,pid,mrp,sp,quantity]
                
        if(productfound==0):
                return cobj.inpid
    else:
        return cobj.cp

@app.route('/add/<pname>/<pid>/<mrp>/<sp>/<quantity>')
def addproduct(pname,pid,mrp,sp,quantity):
    print("endpoinrt started*****************")
    #$code to add product in excel file
    
    if(service.isempty(pname) or service.isempty(pid) or service.isempty(mrp) or service.isempty(sp) or service.isempty(quantity)):
        return "Field is empty"
    elif(service.isemptyzero(mrp) or service.isemptyzero(sp) or service.isemptyzero(quantity)):
        return "data cannot be zero or negative"
    elif(service.isgreater(sp,mrp)):
        return "selling price cannnot be greater than mrp"
    else:
        print("data  assigned*******************")
        try:
            
            status,sheet,wb=config.doConnectionwrite()
            if(status):
                duplicasestatus=service.duplicat(pid)
                print("duplicacy check over******")
                if(duplicasestatus==False):
                    sheet.cell(row=1,column=1).value="product"
                    sheet.cell(row=1,column=2).value="pid"
                    sheet.cell(row=1,column=3).value="mrp"
                    sheet.cell(row=1,column=4).value="sp"
                    sheet.cell(row=1,column=5).value="quantity"
                    print("column created****")
                    
                    #step3 find out no. of data
                    numberofdata=sheet.max_row
                    #step4 insert the data
                    print(numberofdata)
                    sheet.cell(row=numberofdata+1,column=1).value=pname.lower()
                    sheet.cell(row=numberofdata+1,column=2).value=pid
                    print("pnam,pid inserted")
                    sheet.cell(row=numberofdata+1,column=3).value=float(mrp)
                    sheet.cell(row=numberofdata+1,column=4).value=float(sp)
                    sheet.cell(row=numberofdata+1,column=5).value=int(quantity)
                    #step5 save the data
                    print("data ready to******")
                    wb.save('C:\\Users\\lenovo legion 5\\Desktop\\final project\\database\\product.xlsx')
                    print("data saved******")
                    return "ok"
                else:
                    return cobj.dup
        except:
            return False
                
@app.route('/delete')
def deleteproduct():
    return "deleljhuefvbuhejf"

@app.route('/updateByID/<idd>/<newmrp>/<newsp>/<newquantity>')
def updateproductbyid(idd,newmrp,newsp,newquantity):
    
    #step1 -open the database 
    status,sheet=config.doConnectionread()
    if(status):
        #step3 -find out total no of data present in server
        row=sheet.nrows
        print(row)
        #step 4 -retrieve the product details
        productfound=0
        requiredindex=0
        for i in range(row):
            pid=sheet.cell_value(i,1)
            if(idd==pid):
                productfound=1
                requiredindex=i
                oldquantity=sheet.cell_value(i,4)
                
        if(productfound==1):
            if(service.isempty(newmrp) or service.isempty(newsp) or service.isempty(newquantity)):
                return "Field is empty"
            elif(service.isemptyzero(newmrp) or service.isemptyzero(newsp) or service.isemptyzero(newquantity)):
                return "data cannot be zero or negative"
            elif(service.isgreater(newsp,newmrp)):
                return "selling price cannnot be greater than mrp"
            else:
                totalquantity=int(oldquantity)+int(newquantity)
                v=service.updatebyid2(idd,newmrp,newsp,totalquantity,requiredindex)
                print(v)
                return cobj.ok
            
            
        else:
            return cobj.NotOK
                
    else:
        return False
    
@app.route('/updateByname/<newname>/<newmrp>/<newsp>/<newquantity>')
def updateproductbyname(newname,newmrp,newsp,newquantity):
    
    #step1 -open the database 
    status,sheet=config.doConnectionread()
    if(status):
        #step3 -find out total no of data present in server
        row=sheet.nrows
        print(row)
        #step 4 -retrieve the product details
        productfound=0
        requiredindex=0
        for i in range(row):
            pname=sheet.cell_value(i,0)
            if(newname==pname):
                productfound=1
                requiredindex=i
                oldquantity=sheet.cell_value(i,4)
                
        if(productfound==1):
            if(service.isempty(newmrp) or service.isempty(newsp) or service.isempty(newquantity)):
                return "Field is empty"
            elif(service.isemptyzero(newmrp) or service.isemptyzero(newsp) or service.isemptyzero(newquantity)):
                return "data cannot be zero or negative"
            elif(service.isgreater(newsp,newmrp)):
                return "selling price cannnot be greater than mrp"
            else:
                
                
                totalquantity=int(oldquantity)+int(newquantity)
                v=service.updatebyname2(newname,newmrp,newsp,totalquantity,requiredindex)
                print(v)
                return cobj.ok
            
            
        else:
            return cobj.NotOK
                
    else:
        return False 
    
@app.route('/getstockbyname/<getname>')
def getstockbyname(getname):
    status,sheet=config.doConnectionread()
    if(status):
        row=sheet.nrows
        print(row)
        productfound=0
        requiredindex=0
        for i in range(row):
            print(i)
            pname=sheet.cell_value(i,0)
            print(pname,getname)
            if(getname==pname):
                productfound=1
                requiredindex=i
                quantity=sheet.cell_value(i,4)
        if(productfound==1):
            return str(int(quantity))
        else:
            return cobj.NotOK
        
        
            
    else:
        return False
    
@app.route('/getstockbyid/<getid>')
def getstockbyid(getid):
    status,sheet=config.doConnectionread()
    if(status):
        row=sheet.nrows
        print(row)
        productfound=0
        requiredindex=0
        for i in range(row):
            print(i)
            pid=sheet.cell_value(i,1)
            print(pid,getid)
            if(getid==pid):
                productfound=1
                requiredindex=i
                quantity=sheet.cell_value(i,4)
        if(productfound==1):
            return str(int(quantity))
        else:
            return cobj.NotOK
        
        
            
    else:
        return False
    
@app.route('/generatebill/<idd>/<quantity>')
def generatebill(idd,quantity):
    status,sheet=config.doConnectionread()
    if(status):
        row=sheet.nrows
        print(row)
        productfound=0
        for i in range(row):
            pid=sheet.cell_value(i,1)
            if(idd==pid):
                productfound=1
                pname=sheet.cell_value(i,0)
                mrp=sheet.cell_value(i,2)
                sp=sheet.cell_value(i,3)
                
                
        
        if(productfound==0):
            return cobj.NotOK
        else:
            sq=getstockbyid(idd)
            if(int(quantity)<=int(sq)):
                pcart.append(pname)
                spcart.append(sp)
                mrpcart.append(mrp)
                qcart.append(quantity)
                cart["product"]=pcart
                cart["SellingPrice"]=spcart
                cart["MRP"]=mrpcart
                cart["Quantity"]=qcart
                
                return str(cart)
                
               
            else:
                return cobj.os
            
            
            
       
        
    else:
        return cobj.cp
  
    
@app.route('/checkout')
def checkout():
    status,sheet,wb=config.docustomerwrite()
    if(status):
        print("checkout proceed.........................................................")
        allp=cart.get('product')
        allsp=cart.get('SellingPrice')
        allmrp=cart.get('MRP')
        allquantity=cart.get('Quantity')
        tsp=0
        tmrp=0
        tp=len(allp)
        customerid=str(random.randint(10000, 99990))
        discount=0
        for i in range(len(allp)):
            tsp=tsp+(int(allsp[i])*int(allquantity[i]))
            tmrp=tmrp+(int(allmrp[i])*int(allquantity[i]))
        discount=tmrp-tsp
        discoutpercentage=(discount/tmrp)*100
        
        
        customerbill['product']=allp
        customerbill['sellingprice']=tsp
        customerbill['mrp']=tmrp
        customerbill['discount']=discount
        customerbill['DiscountPercentage']=discoutpercentage
        customerbill['Quantity']=allquantity
        customerbill['Total Prodoct']=tp
        customerbill['customerid']=customerid
        print("chkout end*******************************************") 
        return str(customerbill)
    else:
        return cobj.NotOK
     
@app.route('/paylater')
def paylater():
    #fetch the detrials from dis
    p=customerbill.get('product')
    sp=customerbill.get('sellingprice')
    mrp=customerbill.get('mrp')
    quantity=customerbill.get('Quantity')
    discount=customerbill.get('discount')
    discountpercentage=customerbill.get('DiscountPercentage')
    tp=customerbill.get('Total Prodoct')
    cid=customerbill.get('customerid')
    try:
        
        status,sheet,wb=config.docustomerwrite()
        print("paylater proceed.........................................................")
        if(status):
           
                sheet.cell(row=1,column=1).value="product"
                sheet.cell(row=1,column=2).value="discount"
                sheet.cell(row=1,column=3).value="mrp"
                sheet.cell(row=1,column=4).value="sp"
                sheet.cell(row=1,column=5).value="quantity"
                sheet.cell(row=1,column=6).value="discount percentag"
                sheet.cell(row=1,column=7).value="Total product"
                sheet.cell(row=1,column=8).value="Customer ID"
                sheet.cell(row=1,column=9).value="status"
                print("column created**********************************************")
                
                #step3 find out no. of data
                numberofdata=sheet.max_row
                #step4 insert the data
                print(numberofdata)
                sheet.cell(row=numberofdata+1,column=1).value=str(p)
                sheet.cell(row=numberofdata+1,column=2).value=str(discount)
                sheet.cell(row=numberofdata+1,column=3).value=str(mrp)
                sheet.cell(row=numberofdata+1,column=4).value=str(sp)
                sheet.cell(row=numberofdata+1,column=5).value=str(quantity)
                sheet.cell(row=numberofdata+1,column=6).value=str(discountpercentage)
                sheet.cell(row=numberofdata+1,column=7).value=str(tp)
                sheet.cell(row=numberofdata+1,column=8).value=str(cid)
                sheet.cell(row=numberofdata+1,column=9).value="True"

                #step5 save the data
                print("data ready to******")
                wb.save('C:\\Users\\lenovo legion 5\\Desktop\\final project\\database\\paylater.xlsx')
                print("data saved******")
                return cobj.ok
        else:
            return cobj.NotOK
           
    except:
        return cobj.NotOK

'''    
@app.route('/fetchpaylaterdetails')
def fetchpaylaterdetails():
    status,sheet=config.docustomeread()
    if(status):
        cidd=[]
        row=sheet.nrows
        print(row)
        for i in range(row):
            cid=sheet.cell_value(i,7)
            cidd.append(cid)
        return str(cidd)
    else:
        return cobj.cp'''
    
@app.route('/ciddetails/<id>')
def ciddetails(id):
    status,sheet=config.docustomeread()
    if(status):
        row=sheet.nrows
        print(row)
        productfound=0
        
        
        for i in range(row):
            cid=sheet.cell_value(i,7)
            pstatus=sheet.cell_value(i,8)
            if(id==cid):
                productfound=1
                if(pstatus=="True"):
                    p=sheet.cell_value(i,0)
                    discount=sheet.cell_value(i,1)
                    mrp=sheet.cell_value(i,2)
                    sp=sheet.cell_value(i,3)
                    quantity=sheet.cell_value(i,4)
                    discountpercentage=sheet.cell_value(i,5)
                    tp=sheet.cell_value(i,6)
                    cid=sheet.cell_value(i,7)
                    dict={"productname":p,"mrp":mrp,"sp":sp,"quantity":quantity,"discount":discount,"discountpercentage":discountpercentage,"totalproduct":tp}
                    return dict
                    
                else:
                    return cobj.p
        if(productfound==0):
            return cobj.NotOK
    else:
        return cobj.cp
 
def update(ri):
    status,sheet,wb=config.docustomerwrite()
    print("update called***************************************")
    if(status):
        sheet.cell(row=ri+1,column=9).value="False"
        print('saving changes********************************')
        wb.save('C:\\Users\\lenovo legion 5\\Desktop\\final project\\database\\paylater.xlsx')
        print("saved")
        update2(ri)
        return cobj.ok
    else:
        return cobj.cp
    
def update2(ri):
    status,sheet=config.docustomeread()
    print("update called***************************************")
    if(status):
        print('saving changes********************************')
        p=sheet.cell_value(ri,0)
        discount=sheet.cell_value(ri,1)
        mrp=sheet.cell_value(ri,2)
        sp=sheet.cell_value(ri,3)
        quantity=sheet.cell_value(ri,4)
        discountpercentage=sheet.cell_value(ri,5)
        tp=sheet.cell_value(ri,6)
        cid=sheet.cell_value(ri,7)
        dictt={"productname":p,"discount":discount,"mrp":mrp,"sp":sp,"quantity":quantity,"discountpercentage":discountpercentage,"tp":tp,"cid":cid}
        update3(dictt)
        return dictt
    else:
        return cobj.cp
    
    
def update3(dictt):
    p=dictt["productname"]
    discount=dictt["discount"]
    mrp=dictt["mrp"]
    sp=dictt["sp"]
    quantity=dictt["quantity"]
    discountpercentage=dictt["discountpercentage"]
    tp=dictt["tp"]
    cid=dictt["cid"]
    status,sheet,wb=config.dopaymentwrite()
    print("update called***************************************")
    if(status): 
        numberofdata=sheet.max_row
        sheet.cell(row=1,column=1).value="product"
        sheet.cell(row=1,column=2).value="discount"
        sheet.cell(row=1,column=3).value="mrp"
        sheet.cell(row=1,column=4).value="sp"
        sheet.cell(row=1,column=5).value="quantity"
        sheet.cell(row=1,column=6).value="discount percentag"
        sheet.cell(row=1,column=7).value="Total product"
        sheet.cell(row=1,column=8).value="Customer ID"
        
        sheet.cell(row=numberofdata+1,column=1).value=str(p)
        sheet.cell(row=numberofdata+1,column=2).value=str(discount)
        sheet.cell(row=numberofdata+1,column=3).value=str(mrp)
        sheet.cell(row=numberofdata+1,column=4).value=str(sp)
        sheet.cell(row=numberofdata+1,column=5).value=str(quantity)
        sheet.cell(row=numberofdata+1,column=6).value=str(discountpercentage)
        sheet.cell(row=numberofdata+1,column=7).value=str(tp)
        sheet.cell(row=numberofdata+1,column=8).value=str(cid)
        wb.save('C:\\Users\\lenovo legion 5\\Desktop\\final project\\database\\payment.xlsx')
        
    else:
        return cobj.cp
             
@app.route('/confirmdetails/<idd>')
def confirmdetails(idd):
    status,sheet=config.docustomeread()
    if(status):
        row=sheet.nrows
        print(row)
        productfound=0
        for i in range(row):
            cid=sheet.cell_value(i,7)
            ri=0
            if(idd==cid):
                productfound=1
                ri=i
                print("ri found")
                update(ri)
                return cobj.ok
        else:
            return cobj.NotOK
    else:
        return cobj.cp
    
@app.route('/pendingid')
def pendingid():
    status,sheet=config.docustomeread()
    if(status):
        row=sheet.nrows
        print(row)
        productfound=0
        l=[]
        for i in range(row):
            pstatus=sheet.cell_value(i,8)
            if(pstatus=="True"):
                productfound=1
                ri=i
                cid=sheet.cell_value(i,7)
                l.append(cid)
        return l
        if(productfound==0):
            return cobj.NotOK
    else:
        cobj.cp
        
@app.route('/continuepayment')
def continuepayment():
    p=customerbill.get('product')
    sp=customerbill.get('sellingprice')
    mrp=customerbill.get('mrp')
    quantity=customerbill.get('Quantity')
    discount=customerbill.get('discount')
    discountpercentage=customerbill.get('DiscountPercentage')
    tp=customerbill.get('Total Prodoct')
    cid=customerbill.get('customerid')
    try:
        
        status,sheet,wb=config.docustomerwrite()
        print("checkout proceed.........................................................")
        if(status):
           
                sheet.cell(row=1,column=1).value="product"
                sheet.cell(row=1,column=2).value="discount"
                sheet.cell(row=1,column=3).value="mrp"
                sheet.cell(row=1,column=4).value="sp"
                sheet.cell(row=1,column=5).value="quantity"
                sheet.cell(row=1,column=6).value="discount percentag"
                sheet.cell(row=1,column=7).value="Total product"
                sheet.cell(row=1,column=8).value="Customer ID"
                print("column created**********************************************")
                
                #step3 find out no. of data
                numberofdata=sheet.max_row
                #step4 insert the data
                print(numberofdata)
                sheet.cell(row=numberofdata+1,column=1).value=str(p)
                sheet.cell(row=numberofdata+1,column=2).value=str(discount)
                sheet.cell(row=numberofdata+1,column=3).value=str(mrp)
                sheet.cell(row=numberofdata+1,column=4).value=str(sp)
                sheet.cell(row=numberofdata+1,column=5).value=str(quantity)
                sheet.cell(row=numberofdata+1,column=6).value=str(discountpercentage)
                sheet.cell(row=numberofdata+1,column=7).value=str(tp)
                sheet.cell(row=numberofdata+1,column=8).value=str(cid)

                #step5 save the data
                print("data ready to******")
                wb.save('C:\\Users\\lenovo legion 5\\Desktop\\final project\\database\\payment.xlsx')
                print("data saved******")
                return cobj.ok
        else:
            return cobj.NotOK
           
    except:
        return cobj.NotOK
   

if __name__=="__main__":
    app.run(debug=True)

