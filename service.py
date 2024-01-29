from openpyxl import *
from connection import config

def duplicat(pid):
    status,sheet=config.doConnectionread()
    productfound=0
    print("dupicacy connection done")
    row=sheet.nrows
    for i in range(row):
        pidd=sheet.cell_value(i,1)
        print(pidd)
        #print(dbuserid)
        if(pidd==pid):
            productfound=1
          
    if(productfound==1):
        return True
    else:
        return False
    
    
def isempty(para):
    if(para==""):
        return True
    else:
        return False
  
def isemptyzero(para):
    if(para=="0" or int(para)<0):
        return True
    else:
        return False
    
def isgreater(sp,mrp):
    if(int(sp)>int(mrp)):
        return True
    else:
        return False
    
    
def updatebyid2(idd,newmrp,newsp,totalquantity,requiredindex):
    status,sheet,wb=config.doConnectionwrite()
    print("connection successful")
    if(status):
        print("status Ok")
        sheet.cell(row=requiredindex+1,column=3).value=int(newmrp)
        sheet.cell(row=requiredindex+1,column=4).value=int(newsp)
        
        sheet.cell(row=requiredindex+1,column=5).value=int(totalquantity)
        print("product added")
        #save the data
        wb.save('C:\\Users\\lenovo legion 5\\Desktop\\final project\\database\\product.xlsx')
        print("data updated")
        return True
    else:
        return False
    
def updatebyname2(newname,newmrp,newsp,totalquantity,requiredindex):
    status,sheet,wb=config.doConnectionwrite()
    print("connection successful")
    if(status):
        print("status Ok")
        sheet.cell(row=requiredindex+1,column=3).value=int(newmrp)
        sheet.cell(row=requiredindex+1,column=4).value=int(newsp)
        sheet.cell(row=requiredindex+1,column=5).value=int(totalquantity)
        print("product added")
        #save the data
        wb.save('C:\\Users\\lenovo legion 5\\Desktop\\final project\\database\\product.xlsx')
        print("data updated")
        return True
    else:
        return False
    
    
